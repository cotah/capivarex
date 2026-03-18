# -*- coding: utf-8 -*-
"""
Mercado Service
===============
Lista de compras inteligente + análise de notas fiscais com IA.

Features:
- Lista de compras via Supabase (multi-tenant, cada user tem a sua)
- Processamento de notas fiscais (GPT-4 Vision + Google Vision fallback)
- Histórico de preços por produto e mercado no Supabase
- Relatórios mensais de gastos
- Comparação de preços entre mercados
- Alertas automáticos de subida de preço (>20%)
"""

import base64
import logging
from datetime import date, datetime
from typing import Any, Dict, List, Optional

import httpx

from services.core import (
    BaseService,
    register_service,
    ServiceUnavailableError,
)
from services.ai.model_config import INTENT_MODEL, VISION_MODEL
from services.business.grocery_synonyms import find_synonym_in_list
from services.i18n import t

logger = logging.getLogger(__name__)

# ── Constantes ─────────────────────────────────────────────────
PRICE_ALERT_THRESHOLD = 0.20  # 20% de subida dispara alerta

# ── Categorias automáticas ─────────────────────────────────────
_CATEGORIAS: Dict[str, List[str]] = {
    "Laticínios": [
        "leite",
        "manteiga",
        "queijo",
        "iogurte",
        "natas",
        "requeijão",
    ],
    "Padaria": [
        "pão",
        "bolo",
        "croissant",
        "broa",
        "torrada",
    ],
    "Carnes": [
        "frango",
        "carne",
        "bife",
        "peru",
        "porco",
        "vaca",
        "alheira",
    ],
    "Peixe": [
        "peixe",
        "bacalhau",
        "atum",
        "sardinha",
        "salmão",
        "camarão",
    ],
    "Frutas": [
        "maçã",
        "banana",
        "laranja",
        "uva",
        "morango",
        "pera",
        "limão",
    ],
    "Legumes": [
        "alface",
        "tomate",
        "cenoura",
        "cebola",
        "batata",
        "alho",
    ],
    "Bebidas": [
        "água",
        "sumo",
        "cerveja",
        "vinho",
        "refrigerante",
        "café",
    ],
    "Limpeza": [
        "detergente",
        "sabão",
        "lixívia",
        "amaciador",
    ],
    "Higiene": [
        "shampoo",
        "gel",
        "pasta",
        "desodorizante",
        "papel higiénico",
    ],
    "Mercearia": [
        "arroz",
        "massa",
        "feijão",
        "lentilha",
        "grão",
        "açúcar",
        "farinha",
        "óleo",
        "azeite",
    ],
    "Ovos": ["ovo", "ovos"],
    "Snacks": [
        "bolacha",
        "chips",
        "chocolate",
        "biscoito",
    ],
}

_CAT_I18N = {
    "Laticínios": "cat_dairy",
    "Padaria": "cat_bakery",
    "Carnes": "cat_meat",
    "Peixe": "cat_fish",
    "Frutas": "cat_fruit",
    "Legumes": "cat_vegetables",
    "Bebidas": "cat_drinks",
    "Limpeza": "cat_cleaning",
    "Higiene": "cat_hygiene",
    "Mercearia": "cat_grocery",
    "Ovos": "cat_eggs",
    "Snacks": "cat_snacks",
    "Outros": "cat_other",
}

_VISION_PROMPT = (
    "You are an expert receipt/invoice OCR system. "
    "Analyze this supermarket receipt image and extract ALL purchased items.\n\n"
    "CRITICAL RULES for accurate extraction:\n"
    "1. Each item's price must come from the SAME LINE as the product name.\n"
    "2. Do NOT assign a price from an adjacent line to a different product.\n"
    "3. If a product spans multiple lines, its price is on the LAST line of that product.\n"
    "4. Discount lines (negative values) should be IGNORED — do not create items for discounts.\n"
    "5. The sum of all preco_total values MUST equal the receipt total (within €0.05).\n"
    "6. If you cannot confidently read a price, set it to null rather than guessing.\n\n"
    "Respond ONLY with valid JSON, no markdown, no explanations:\n\n"
    "{\n"
    '  "mercado": "Store name (e.g. Lidl, Aldi, Tesco, Dunnes)",\n'
    '  "data": "DD/MM/YYYY or null if not visible",\n'
    '  "total": 0.00,\n'
    '  "itens": [\n'
    "    {\n"
    '      "produto": "Product name normalized (keep original language)",\n'
    '      "quantidade": 1,\n'
    '      "unidade": "un/kg/L",\n'
    '      "preco_total": 0.00,\n'
    '      "preco_unitario": 0.00\n'
    "    }\n"
    "  ]\n"
    "}\n\n"
    "Additional rules:\n"
    "- Normalize abbreviated names: 'LT MIMOSA MEIO-GORD' → 'Leite Mimosa Meio-Gordo'\n"
    "- preco_unitario = preco_total / quantidade\n"
    "- Include ALL items, even partially readable ones\n"
    "- VERIFY: sum of all preco_total should match the total on the receipt"
)


def _categorizar(produto: str) -> str:
    nome = produto.lower()
    for categoria, keywords in _CATEGORIAS.items():
        if any(kw in nome for kw in keywords):
            return categoria
    return "Outros"


@register_service("mercado")
class MercadoService(BaseService):
    """
    Mercado inteligente: lista de compras + notas
    fiscais + análise de preços.
    Agora usa Supabase directamente para lista de
    compras (sem N8N).
    """

    def __init__(self):
        super().__init__(name="mercado")
        self._http: Optional[httpx.AsyncClient] = None
        self._db = None

    async def _initialize(self) -> None:
        self._http = httpx.AsyncClient(timeout=20.0)
        from services import get_service

        self._db = get_service("database")
        self.logger.info("MercadoService initialized (Supabase mode)")

    async def _health_check(self) -> bool:
        return self._db is not None and self._db.is_initialized()

    def _get_client(self):
        """Retorna o cliente Supabase."""
        if not self._db or not self._db.is_initialized():
            raise ServiceUnavailableError("Database service not available")
        return self._db.get_client()

    # ── Lista de compras (via Supabase) ────────────────

    async def adicionar(
        self,
        itens: str,
        user_id: str = "default",
        lang: str = "en",
    ) -> Dict[str, Any]:
        """Adiciona item(s) à lista de compras."""
        import asyncio
        import re

        itens_list = re.split(r"[,;\n]|\se\s", itens)
        itens_list = [i.strip() for i in itens_list if i.strip()]

        if not itens_list:
            return {"mensagem": t("mercado_what_to_add", lang=lang)}

        db = self._get_client()
        adicionados = []
        duplicados = []

        # Fetch existing list once for synonym check
        lista_existente = await self._get_lista(user_id)
        existing_lower = {i.lower() for i in lista_existente}

        for item in itens_list:
            item_clean = item.strip()
            item_lower = item_clean.lower()

            # Check exact match first (fast)
            if item_lower in existing_lower:
                duplicados.append(item_clean)
                continue

            # Check cross-language synonym match
            synonym_match = find_synonym_in_list(item_clean, lista_existente)
            if synonym_match:
                duplicados.append(f"{item_clean} (= {synonym_match})")
                continue

            try:
                await asyncio.get_event_loop().run_in_executor(
                    None,
                    lambda i=item_clean: (
                        db.table("shopping_list")
                        .insert(
                            {
                                "user_id": user_id,
                                "item": i,
                                "status": "pendente",
                            }
                        )
                        .execute()
                    ),
                )
                adicionados.append(item_clean)
                # Update local state so subsequent items in same batch
                # can detect duplicates/synonyms
                lista_existente.append(item_clean)
                existing_lower.add(item_lower)
            except Exception as e:
                if (
                    "idx_shopping_list_unique" in str(e)
                    or "duplicate" in str(e).lower()
                ):
                    duplicados.append(item_clean)
                else:
                    self.logger.warning(
                        "Erro ao adicionar '%s': %s",
                        item_clean,
                        e,
                    )
                    duplicados.append(item_clean)

        lista = await self._get_lista(user_id)

        mensagem = ""
        if adicionados:
            mensagem += t("mercado_added", lang=lang, items=", ".join(adicionados))
        if duplicados:
            mensagem += t(
                "mercado_already_in_list", lang=lang, items=", ".join(duplicados)
            )
        mensagem += t("mercado_current_list", lang=lang, total=len(lista))
        mensagem += "\n".join(f"{i + 1}. {item}" for i, item in enumerate(lista))

        return {
            "action": "adicionar",
            "adicionados": adicionados,
            "duplicados": duplicados,
            "lista": lista,
            "total": len(lista),
            "mensagem": mensagem,
        }

    async def ver_lista(
        self, user_id: str = "default", lang: str = "en"
    ) -> Dict[str, Any]:
        """Mostra a lista de compras actual."""
        lista = await self._get_lista(user_id)

        if not lista:
            mensagem = t("mercado_list_empty", lang=lang)
        else:
            mensagem = t("mercado_list_header", lang=lang, total=len(lista))
            mensagem += "\n".join(f"{i + 1}. {item}" for i, item in enumerate(lista))
            mensagem += t("mercado_list_footer", lang=lang)

        return {
            "action": "lista",
            "lista": lista,
            "total": len(lista),
            "mensagem": mensagem,
        }

    async def remover(
        self,
        item: str,
        user_id: str = "default",
        lang: str = "en",
    ) -> Dict[str, Any]:
        """Remove um item da lista."""
        import asyncio

        if not item:
            return {
                "action": "remover",
                "erro": True,
                "mensagem": t("mercado_what_to_remove", lang=lang),
            }

        db = self._get_client()

        try:
            result = await asyncio.get_event_loop().run_in_executor(
                None,
                lambda: (
                    db.table("shopping_list")
                    .delete()
                    .eq("user_id", user_id)
                    .eq("status", "pendente")
                    .ilike("item", item)
                    .execute()
                ),
            )
            removeu = bool(result.data)
        except Exception as e:
            self.logger.warning("Erro ao remover '%s': %s", item, e)
            removeu = False

        lista = await self._get_lista(user_id)

        if removeu:
            mensagem = t("mercado_item_removed", lang=lang, item=item, total=len(lista))
        else:
            mensagem = t("mercado_item_not_found", lang=lang, item=item)

        return {
            "action": "remover",
            "item": item,
            "removeu": removeu,
            "lista": lista,
            "total": len(lista),
            "mensagem": mensagem,
        }

    async def limpar_lista(
        self, user_id: str = "default", lang: str = "en"
    ) -> Dict[str, Any]:
        """Limpa a lista toda e guarda no histórico."""
        import asyncio

        lista = await self._get_lista(user_id)
        total_antes = len(lista)

        db = self._get_client()

        # Guardar no histórico antes de limpar
        if lista:
            try:
                await asyncio.get_event_loop().run_in_executor(
                    None,
                    lambda: (
                        db.table("shopping_list_history")
                        .insert(
                            {
                                "user_id": user_id,
                                "items": lista,
                            }
                        )
                        .execute()
                    ),
                )
            except Exception as e:
                self.logger.warning("Erro ao guardar histórico: %s", e)

        # Limpar lista
        try:
            await asyncio.get_event_loop().run_in_executor(
                None,
                lambda: (
                    db.table("shopping_list")
                    .delete()
                    .eq("user_id", user_id)
                    .eq("status", "pendente")
                    .execute()
                ),
            )
        except Exception as e:
            self.logger.error("Erro ao limpar lista: %s", e)
            return {"mensagem": t("mercado_clear_error", lang=lang)}

        return {
            "action": "limpar",
            "totalRemovido": total_antes,
            "lista": [],
            "mensagem": t("mercado_list_cleared", lang=lang, total=total_antes),
        }

    async def historico_lista(
        self, user_id: str = "default", lang: str = "en"
    ) -> Dict[str, Any]:
        """Mostra o histórico de listas limpas."""
        import asyncio

        db = self._get_client()

        try:
            result = await asyncio.get_event_loop().run_in_executor(
                None,
                lambda: (
                    db.table("shopping_list_history")
                    .select("items, cleared_at")
                    .eq("user_id", user_id)
                    .order("cleared_at", desc=True)
                    .limit(5)
                    .execute()
                ),
            )
            historico = result.data or []
        except Exception as e:
            self.logger.warning("Erro ao buscar histórico: %s", e)
            historico = []

        if not historico:
            mensagem = t("mercado_no_history", lang=lang)
        else:
            mensagem = t("mercado_history_header", lang=lang)
            for i, h in enumerate(historico):
                data = h.get("cleared_at", "")[:10]
                items = h.get("items", [])
                if isinstance(items, list):
                    items_str = ", ".join(items[:10])
                    if len(items) > 10:
                        items_str += f" (+{len(items) - 10})"
                else:
                    items_str = str(items)
                mensagem += f"{i + 1}. {data}: {items_str}\n"

        return {
            "action": "historico",
            "historico": historico,
            "mensagem": mensagem,
        }

    async def _get_lista(self, user_id: str = "default") -> List[str]:
        """Helper: retorna lista de itens pendentes."""
        import asyncio

        db = self._get_client()
        try:
            result = await asyncio.get_event_loop().run_in_executor(
                None,
                lambda: (
                    db.table("shopping_list")
                    .select("item")
                    .eq("user_id", user_id)
                    .eq("status", "pendente")
                    .order("created_at")
                    .execute()
                ),
            )
            return [r["item"] for r in (result.data or [])]
        except Exception as e:
            self.logger.warning("Erro ao buscar lista: %s", e)
            return []

    # ═══════════════════════════════════════════════════
    # TUDO ABAIXO FICA EXACTAMENTE IGUAL — SEM ALTERAÇÕES
    # ═══════════════════════════════════════════════════

    # ── Processamento de nota fiscal ───────────────────

    async def processar_nota(
        self,
        image_data: bytes,
        chat_id: str,
        mime_type: str = "image/jpeg",
        lang: str = "en",
    ) -> Dict[str, Any]:
        """
        Processa foto de nota fiscal.
        GPT-4 Vision primeiro, Google Vision como fallback.
        Guarda no Supabase e verifica alertas de preço.
        """
        self.logger.info("Processando nota fiscal chat_id=%s", chat_id)

        # 1) Gemini 2.5 Flash (primario — 16x mais barato)
        nota = await self._extrair_gemini(image_data, mime_type)
        fonte = "gemini_flash"

        # 2) GPT-4o Vision (fallback 1)
        if not nota or not nota.get("itens"):
            self.logger.warning("Gemini OCR falhou, tentando GPT-4 Vision")
            nota = await self._extrair_gpt4(image_data, mime_type)
            fonte = "gpt4_vision"

        # 3) Google Vision OCR (fallback 2)
        if not nota or not nota.get("itens"):
            self.logger.warning("GPT-4 Vision falhou, tentando Google Vision")
            nota = await self._extrair_google_vision(image_data)
            fonte = "google_vision"

        if not nota or not nota.get("itens"):
            return {
                "sucesso": False,
                "mensagem": t("mercado_receipt_failed", lang=lang),
            }

        nota = self._normalizar_nota(nota)
        nota = self._validar_ocr(nota)

        # Log confidence
        if nota.get("ocr_confidence") != "high":
            self.logger.warning(
                "OCR confidence: %s | warnings: %s",
                nota.get("ocr_confidence"),
                nota.get("ocr_warnings"),
            )

        # Verificar duplicata antes de guardar
        duplicada = await self._verificar_duplicata(
            chat_id=chat_id,
            mercado=nota["mercado"],
            data_compra=nota["data_obj"],
            total=float(nota.get("total") or 0),
        )
        if duplicada:
            return {
                "sucesso": True,
                "duplicada": True,
                "mensagem": t(
                    "mercado_duplicate_receipt",
                    lang=lang,
                    mercado=nota["mercado"],
                    total=f"{nota.get('total', 0):.2f}",
                ),
            }

        compra_id = await self._guardar_compra(nota, chat_id, fonte)
        alertas = await self._verificar_alertas(nota["itens"], chat_id, lang=lang)

        return self._formatar_resposta_nota(nota, alertas, compra_id, lang=lang)

    async def _extrair_gemini(
        self, image_data: bytes, mime_type: str
    ) -> Optional[Dict[str, Any]]:
        """Extrai dados da nota via Gemini 2.5 Flash (primario, mais barato)."""
        try:
            import json
            import os

            api_key = os.environ.get("GEMINI_API_KEY")
            if not api_key:
                self.logger.warning("GEMINI_API_KEY not set, skipping Gemini OCR")
                return None

            b64 = base64.b64encode(image_data).decode()

            payload = {
                "contents": [
                    {
                        "parts": [
                            {"text": _VISION_PROMPT},
                            {
                                "inline_data": {
                                    "mime_type": mime_type,
                                    "data": b64,
                                }
                            },
                        ]
                    }
                ],
                "generationConfig": {
                    "temperature": 0.1,
                    "maxOutputTokens": 2000,
                    "responseMimeType": "application/json",
                    "thinkingConfig": {"thinkingBudget": 0},
                },
            }

            resp = await self._http.post(
                "https://generativelanguage.googleapis.com"
                "/v1beta/models/"
                "gemini-2.5-flash"
                ":generateContent"
                f"?key={api_key}",
                json=payload,
                timeout=30.0,
            )

            if resp.status_code != 200:
                self.logger.warning(
                    "Gemini OCR HTTP %d: %s",
                    resp.status_code,
                    resp.text[:200],
                )
                return None

            data = resp.json()
            text = (
                data.get("candidates", [{}])[0]
                .get("content", {})
                .get("parts", [{}])[0]
                .get("text", "")
                .strip()
            )

            if not text:
                return None

            text = text.replace("```json", "").replace("```", "").strip()
            result = json.loads(text)

            self.logger.info(
                "Gemini OCR: %d itens extraidos, mercado=%s",
                len(result.get("itens", [])),
                result.get("mercado", "?"),
            )
            return result

        except Exception as e:
            self.logger.error("Gemini OCR error: %s", e)
            return None

    async def _extrair_gpt4(
        self, image_data: bytes, mime_type: str
    ) -> Optional[Dict[str, Any]]:
        """Extrai dados da nota via GPT-4 Vision."""
        try:
            import json

            from services import get_service

            openai_svc = get_service("openai")
            if not openai_svc or not openai_svc.is_initialized():
                return None

            b64 = base64.b64encode(image_data).decode()
            client = openai_svc.client

            response = await client.chat.completions.create(
                model=VISION_MODEL,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": _VISION_PROMPT,
                            },
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": (f"data:{mime_type};base64,{b64}"),
                                    "detail": "high",
                                },
                            },
                        ],
                    }
                ],
                max_completion_tokens=2000,
                temperature=0.1,
            )

            text = response.choices[0].message.content.strip()
            text = text.replace("```json", "").replace("```", "").strip()
            return json.loads(text)

        except Exception as e:
            self.logger.error("GPT-4 Vision error: %s", e)
            return None

    async def _extrair_google_vision(
        self, image_data: bytes
    ) -> Optional[Dict[str, Any]]:
        """Fallback: Google Vision OCR + GPT-4 text."""
        try:
            import json

            b64 = base64.b64encode(image_data).decode()
            resp = await self._http.post(
                "https://vision.googleapis.com/v1/images:annotate",
                json={
                    "requests": [
                        {
                            "image": {"content": b64},
                            "features": [{"type": "TEXT_DETECTION"}],
                        }
                    ]
                },
            )
            if resp.status_code != 200:
                return None

            text_ocr = (
                resp.json()
                .get("responses", [{}])[0]
                .get("textAnnotations", [{}])[0]
                .get("description", "")
            )
            if not text_ocr or len(text_ocr) < 20:
                return None

            from services import get_service

            openai_svc = get_service("openai")
            if not openai_svc or not openai_svc.is_initialized():
                return None

            prompt = f"{_VISION_PROMPT}\n\nTexto OCR extraído:\n{text_ocr}"
            response = await openai_svc.client.chat.completions.create(
                model=INTENT_MODEL,
                messages=[
                    {
                        "role": "user",
                        "content": prompt,
                    }
                ],
                max_completion_tokens=2000,
                temperature=0.1,
            )
            raw = response.choices[0].message.content.strip()
            raw = raw.replace("```json", "").replace("```", "").strip()
            return json.loads(raw)

        except Exception as e:
            self.logger.error("Google Vision fallback error: %s", e)
            return None

    def _normalizar_nota(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Normaliza dados extraídos."""
        data_raw = data.get("data")
        data_obj = date.today()
        if data_raw:
            for fmt in (
                "%d/%m/%Y",
                "%d/%m/%y",
                "%Y-%m-%d",
            ):
                try:
                    data_obj = datetime.strptime(data_raw, fmt).date()
                    break
                except ValueError:
                    continue
        data["data_obj"] = data_obj

        if not data.get("total"):
            data["total"] = round(
                sum(float(i.get("preco_total") or 0) for i in data.get("itens", [])),
                2,
            )

        if not data.get("mercado"):
            data["mercado"] = "Desconhecido"

        itens_ok = []
        for item in data.get("itens", []):
            if not item.get("produto"):
                continue
            qtd = float(item.get("quantidade") or 1)
            p_total = float(item.get("preco_total") or 0)
            p_unit = float(item.get("preco_unitario") or 0)
            if p_unit == 0 and p_total > 0 and qtd > 0:
                p_unit = round(p_total / qtd, 4)
            if p_total == 0 and p_unit > 0:
                p_total = round(p_unit * qtd, 2)
            item.update(
                {
                    "preco_total": p_total,
                    "preco_unitario": p_unit,
                    "quantidade": qtd,
                    "unidade": (item.get("unidade") or "un"),
                    "categoria": _categorizar(item["produto"]),
                }
            )
            itens_ok.append(item)
        data["itens"] = itens_ok
        return data

    def _validar_ocr(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Post-OCR validation and correction.

        Checks:
        1. Sum of items vs receipt total (tolerance €0.10)
        2. Items with null/zero prices
        3. Suspicious duplicate prices in adjacent items
        4. Individual item price > total
        5. Negative prices (discount lines that slipped through)

        Returns:
            Updated data dict with:
            - 'ocr_warnings': list of warning strings
            - 'ocr_confidence': 'high' | 'medium' | 'low'
            - Items corrected where possible
        """
        warnings: List[str] = []
        itens = data.get("itens", [])
        total_receipt = float(data.get("total") or 0)

        # ── Check 1: Sum vs total ────────────────────────────
        soma_itens = sum(float(i.get("preco_total") or 0) for i in itens)

        if total_receipt > 0 and abs(soma_itens - total_receipt) > 0.10:
            diff = abs(soma_itens - total_receipt)
            warnings.append(
                f"Sum mismatch: items={soma_itens:.2f}, "
                f"receipt={total_receipt:.2f}, diff={diff:.2f}"
            )

            # If sum is way off but total exists, trust the total
            # and flag the data as low confidence
            if diff > total_receipt * 0.15:  # >15% off
                data["ocr_confidence"] = "low"
            else:
                data["ocr_confidence"] = "medium"

        # ── Check 2: Null/zero prices ────────────────────────
        null_prices = [i for i in itens if not i.get("preco_total")]
        if null_prices:
            warnings.append(f"{len(null_prices)} item(s) with no price")
            # Try to calculate from total if only 1 item missing
            if len(null_prices) == 1 and total_receipt > 0:
                known_sum = sum(
                    float(i.get("preco_total") or 0)
                    for i in itens
                    if i.get("preco_total")
                )
                missing_price = round(total_receipt - known_sum, 2)
                if missing_price > 0:
                    null_prices[0]["preco_total"] = missing_price
                    null_prices[0]["preco_unitario"] = round(
                        missing_price / float(null_prices[0].get("quantidade") or 1),
                        4,
                    )
                    warnings.append(
                        f"Inferred price for "
                        f"'{null_prices[0]['produto']}': €{missing_price:.2f}"
                    )

        # ── Check 3: Duplicate adjacent prices ───────────────
        for i in range(len(itens) - 1):
            p1 = float(itens[i].get("preco_total") or 0)
            p2 = float(itens[i + 1].get("preco_total") or 0)
            if p1 > 0 and p1 == p2:
                # Same price for adjacent items is suspicious
                # (could be legitimate but worth flagging)
                warnings.append(
                    f"Duplicate adjacent price €{p1:.2f}: "
                    f"'{itens[i]['produto']}' and '{itens[i + 1]['produto']}'"
                )

        # ── Check 4: Price > total ───────────────────────────
        if total_receipt > 0:
            for item in itens:
                p = float(item.get("preco_total") or 0)
                if p > total_receipt:
                    warnings.append(
                        f"Item price €{p:.2f} > receipt total "
                        f"€{total_receipt:.2f}: '{item['produto']}'"
                    )
                    # This is clearly wrong — null it out
                    item["preco_total"] = 0
                    item["preco_unitario"] = 0

        # ── Check 5: Negative prices ─────────────────────────
        data["itens"] = [i for i in itens if float(i.get("preco_total") or 0) >= 0]
        removed = len(itens) - len(data["itens"])
        if removed:
            warnings.append(f"Removed {removed} discount/negative line(s)")

        # ── Confidence score ─────────────────────────────────
        if "ocr_confidence" not in data:
            if not warnings:
                data["ocr_confidence"] = "high"
            elif len(warnings) <= 2:
                data["ocr_confidence"] = "medium"
            else:
                data["ocr_confidence"] = "low"

        data["ocr_warnings"] = warnings

        if warnings:
            self.logger.info("OCR validation warnings: %s", "; ".join(warnings))

        return data

    async def _verificar_duplicata(
        self,
        chat_id: str,
        mercado: str,
        data_compra,
        total: float,
    ) -> bool:
        """Verifica se nota já foi registada (mesmo mercado + data + total)."""
        try:
            from services import get_service

            supabase_svc = get_service("database")
            if not supabase_svc or not supabase_svc.is_initialized():
                return False

            db = supabase_svc.client

            result = (
                db.table("mercado_compras")
                .select("id")
                .eq("chat_id", chat_id)
                .eq("mercado", mercado)
                .eq(
                    "data_compra",
                    data_compra.isoformat(),
                )
                .eq("total", total)
                .limit(1)
                .execute()
            )
            return bool(result.data)

        except Exception as e:
            self.logger.warning("Erro ao verificar duplicata: %s", e)
            return False

    async def _guardar_compra(
        self,
        nota: Dict[str, Any],
        chat_id: str,
        fonte: str,
    ) -> Optional[str]:
        """Insere compra e itens no Supabase."""
        try:
            from services import get_service

            supabase_svc = get_service("database")
            if not supabase_svc or not supabase_svc.is_initialized():
                self.logger.warning("Supabase não disponível — compra não guardada")
                return None

            db = supabase_svc.client

            compra_res = (
                db.table("mercado_compras")
                .insert(
                    {
                        "chat_id": chat_id,
                        "mercado": nota["mercado"],
                        "data_compra": (nota["data_obj"].isoformat()),
                        "total": float(nota.get("total") or 0),
                        "fonte_extracao": fonte,
                    }
                )
                .execute()
            )
            compra_id = compra_res.data[0]["id"]

            itens_payload = [
                {
                    "compra_id": compra_id,
                    "chat_id": chat_id,
                    "produto": i["produto"],
                    "categoria": i["categoria"],
                    "quantidade": float(i["quantidade"]),
                    "unidade": i["unidade"],
                    "preco_total": float(i["preco_total"]),
                    "preco_unitario": float(i["preco_unitario"]),
                    "mercado": nota["mercado"],
                    "data_compra": (nota["data_obj"].isoformat()),
                }
                for i in nota["itens"]
                if float(i.get("preco_total") or 0) > 0
            ]

            if itens_payload:
                db.table("mercado_itens").insert(itens_payload).execute()

            self.logger.info(
                "Compra guardada: id=%s, %d itens, €%.2f",
                compra_id,
                len(itens_payload),
                nota.get("total", 0),
            )
            return compra_id

        except Exception as e:
            self.logger.error(
                "Erro ao guardar compra: %s",
                e,
                exc_info=True,
            )
            return None

    async def _verificar_alertas(
        self, itens: List[Dict], chat_id: str, lang: str = "en"
    ) -> List[str]:
        """Detecta subidas de preço > 20%."""
        alertas: List[str] = []
        try:
            from services import get_service

            supabase_svc = get_service("database")
            if not supabase_svc or not supabase_svc.is_initialized():
                return alertas
            db = supabase_svc.client

            for item in itens:
                preco_atual = float(item.get("preco_unitario") or 0)
                if preco_atual <= 0:
                    continue
                res = (
                    db.table("mercado_itens")
                    .select("preco_unitario, mercado, data_compra")
                    .eq("chat_id", chat_id)
                    .ilike(
                        "produto",
                        f"%{item['produto'].split()[0]}%",
                    )
                    .order("data_compra", desc=True)
                    .limit(1)
                    .execute()
                )
                if not res.data:
                    continue
                preco_anterior = float(res.data[0]["preco_unitario"])
                if preco_anterior <= 0:
                    continue
                variacao = (preco_atual - preco_anterior) / preco_anterior
                # Only alert for increases > 20% AND price difference > €0.50
                # (small absolute differences are likely OCR noise)
                diff_abs = abs(preco_atual - preco_anterior)
                if variacao >= PRICE_ALERT_THRESHOLD and diff_abs > 0.50:
                    alertas.append(
                        t(
                            "mercado_price_alert_item",
                            lang=lang,
                            product=item["produto"],
                            pct=f"{variacao * 100:.0f}",
                            old=f"{preco_anterior:.2f}",
                            new=f"{preco_atual:.2f}",
                        )
                    )
        except Exception as e:
            self.logger.warning("Erro ao verificar alertas: %s", e)
        return alertas

    # ------------------------------------------------------------------
    # Proactive Price Drop Detection (weekly check)
    # ------------------------------------------------------------------

    async def verificar_descidas_preco(
        self, chat_id: str, lang: str = "en"
    ) -> Dict[str, Any]:
        """
        Check for significant price drops on user's frequently bought products.

        Compares the latest price of each product with the average of
        previous purchases. Flags drops > 15% AND > €0.30.

        Returns:
            Dict with:
            - has_alerts: bool
            - alerts: List[Dict] with produto, mercado, preco_atual,
              preco_anterior, variacao, economia
            - mensagem: str (formatted message for Telegram)
        """
        try:
            from collections import defaultdict
            from datetime import timedelta

            from services import get_service

            supabase_svc = get_service("database")
            if not supabase_svc or not supabase_svc.is_initialized():
                return {"has_alerts": False, "alerts": [], "mensagem": ""}
            db = supabase_svc.client

            # Get all items from last 90 days for this user
            cutoff = (date.today() - timedelta(days=90)).isoformat()

            res = (
                db.table("mercado_itens")
                .select("produto, preco_unitario, mercado, data_compra")
                .eq("chat_id", chat_id)
                .gte("data_compra", cutoff)
                .order("data_compra", desc=True)
                .execute()
            )

            if not res.data or len(res.data) < 5:
                return {"has_alerts": False, "alerts": [], "mensagem": ""}

            # Group by product: {produto: [{preco, mercado, data}, ...]}
            products: Dict[str, list] = defaultdict(list)
            for item in res.data:
                prod = (item.get("produto") or "").lower().strip()
                preco = float(item.get("preco_unitario") or 0)
                if prod and preco > 0:
                    products[prod].append(
                        {
                            "preco": preco,
                            "mercado": item.get("mercado", ""),
                            "data": item.get("data_compra", ""),
                        }
                    )

            # For each product with 2+ purchases, compare latest vs avg of older
            alerts: List[Dict[str, Any]] = []
            for prod, entries in products.items():
                if len(entries) < 2:
                    continue

                # Latest purchase (already sorted desc by data_compra)
                latest = entries[0]
                preco_atual = latest["preco"]

                # Average of older purchases (skip the latest)
                older_prices = [e["preco"] for e in entries[1:]]
                preco_medio = sum(older_prices) / len(older_prices)

                # Check for significant drop
                if preco_medio > 0 and preco_atual < preco_medio:
                    variacao = ((preco_atual - preco_medio) / preco_medio) * 100
                    economia = round(preco_medio - preco_atual, 2)

                    # Only alert if drop > 15% AND > €0.30
                    if variacao < -15 and economia > 0.30:
                        alerts.append(
                            {
                                "produto": prod.title(),
                                "mercado": latest["mercado"],
                                "preco_atual": preco_atual,
                                "preco_anterior": round(preco_medio, 2),
                                "variacao": round(variacao, 1),
                                "economia": economia,
                            }
                        )

            if not alerts:
                return {"has_alerts": False, "alerts": [], "mensagem": ""}

            # Sort by savings descending
            alerts.sort(key=lambda x: -x["economia"])

            # Build message (max 5 alerts)
            top_alerts = alerts[:5]
            total_economia = sum(a["economia"] for a in top_alerts)

            linhas = [t("mercado_price_drops_header", lang=lang)]
            for a in top_alerts:
                linhas.append(
                    t(
                        "mercado_price_drop_item",
                        lang=lang,
                        product=a["produto"],
                        price=f"{a['preco_atual']:.2f}",
                        store=a["mercado"],
                        old_price=f"{a['preco_anterior']:.2f}",
                        change=f"{a['variacao']:.0f}",
                    )
                )

            linhas.append(
                t(
                    "mercado_price_drops_savings",
                    lang=lang,
                    total=f"{total_economia:.2f}",
                )
            )

            return {
                "has_alerts": True,
                "alerts": top_alerts,
                "mensagem": "\n".join(linhas),
            }

        except Exception as e:
            self.logger.error("Price drop check failed: %s", e, exc_info=True)
            return {"has_alerts": False, "alerts": [], "mensagem": ""}

    def _formatar_resposta_nota(
        self,
        nota: Dict[str, Any],
        alertas: List[str],
        compra_id: Optional[str],
        lang: str = "en",
    ) -> Dict[str, Any]:
        itens = nota["itens"]
        mercado = nota["mercado"]
        data_str = nota["data_obj"].strftime("%d/%m/%Y")
        total = nota.get("total", 0)

        linhas = [
            t("mercado_receipt_header", lang=lang, mercado=mercado),
            t(
                "mercado_receipt_date_summary",
                lang=lang,
                date=data_str,
                count=len(itens),
                total=f"{total:.2f}",
            ),
            "",
            t("mercado_items_header", lang=lang),
        ]
        for item in itens[:15]:
            linha = f" • {item['produto']}"
            if float(item.get("quantidade", 1)) != 1:
                linha += f" ×{item['quantidade']:.0f}{item.get('unidade', '')}"
            linha += f" — €{item.get('preco_total', 0):.2f}"
            linhas.append(linha)
        if len(itens) > 15:
            linhas.append(t("mercado_more_items", lang=lang, count=len(itens) - 15))

        if alertas:
            linhas += [
                "",
                t("mercado_price_alerts_header", lang=lang),
            ] + alertas

        # OCR confidence warnings
        if nota.get("ocr_confidence") == "low":
            linhas.append(t("mercado_ocr_low_confidence", lang=lang))
        elif nota.get("ocr_confidence") == "medium" and nota.get("ocr_warnings"):
            linhas.append(t("mercado_ocr_medium_confidence", lang=lang))

        linhas += [
            "",
            t("mercado_tip_report", lang=lang),
            t("mercado_tip_compare", lang=lang),
        ]

        return {
            "sucesso": True,
            "compra_id": compra_id,
            "mercado": mercado,
            "total": total,
            "n_itens": len(itens),
            "alertas": alertas,
            "mensagem": "\n".join(linhas),
        }

    # ── Relatórios ─────────────────────────────────────

    async def relatorio_mensal(
        self,
        chat_id: str,
        mes: Optional[int] = None,
        ano: Optional[int] = None,
        lang: str = "en",
    ) -> Dict[str, Any]:
        """Relatório de gastos do mês."""
        try:
            from services import get_service

            db = get_service("database").client
            hoje = date.today()
            mes = mes or hoje.month
            ano = ano or hoje.year
            mes_str = f"{ano}-{mes:02d}"

            res = (
                db.table("mercado_compras")
                .select("mercado, total, data_compra")
                .eq("chat_id", chat_id)
                .gte("data_compra", f"{mes_str}-01")
                .lte("data_compra", f"{mes_str}-31")
                .execute()
            )
            compras = res.data or []

            if not compras:
                nome_mes = self._nome_mes(mes, lang=lang)
                return {
                    "mensagem": t(
                        "mercado_no_purchases_month",
                        lang=lang,
                        month=nome_mes,
                        year=ano,
                    )
                }

            por_mercado: Dict[str, Dict] = {}
            total_geral = 0.0
            for c in compras:
                m = c["mercado"]
                v = float(c.get("total") or 0)
                if m not in por_mercado:
                    por_mercado[m] = {
                        "total": 0.0,
                        "visitas": 0,
                    }
                por_mercado[m]["total"] += v
                por_mercado[m]["visitas"] += 1
                total_geral += v

            ranking = sorted(
                por_mercado.items(),
                key=lambda x: x[1]["total"],
                reverse=True,
            )
            nome_mes = self._nome_mes(mes, lang=lang)
            linhas = [
                t("mercado_report_header", lang=lang, month=nome_mes, year=ano),
                t("mercado_total_spent", lang=lang, total=f"{total_geral:.2f}"),
                t("mercado_stores_visited", lang=lang, count=len(por_mercado)),
                "",
                t("mercado_spending_by_store", lang=lang),
            ]
            medalhas = ["🥇", "🥈", "🥉"]
            for i, (mercado, dados) in enumerate(ranking):
                emoji = medalhas[i] if i < 3 else " "
                pct = (dados["total"] / total_geral * 100) if total_geral else 0
                linhas.append(
                    f"{emoji} *{mercado}*: "
                    f"€{dados['total']:.2f} "
                    f"({pct:.0f}%) — "
                    + t("mercado_visit_count", lang=lang, count=dados["visitas"])
                )

            res2 = (
                db.table("mercado_itens")
                .select("categoria, preco_total")
                .eq("chat_id", chat_id)
                .gte("data_compra", f"{mes_str}-01")
                .lte("data_compra", f"{mes_str}-31")
                .execute()
            )
            if res2.data:
                cats: Dict[str, float] = {}
                for item in res2.data:
                    cat = item.get("categoria", "Outros")
                    cats[cat] = cats.get(cat, 0) + float(item.get("preco_total") or 0)
                top = sorted(
                    cats.items(),
                    key=lambda x: x[1],
                    reverse=True,
                )[:5]
                linhas += [
                    "",
                    t("mercado_top_categories", lang=lang),
                ]
                for cat, val in top:
                    cat_key = _CAT_I18N.get(cat, "cat_other")
                    linhas.append(f" • {t(cat_key, lang=lang)}: €{val:.2f}")

            return {
                "mensagem": "\n".join(linhas),
                "total": total_geral,
            }

        except Exception as e:
            self.logger.error(
                "Erro no relatório mensal: %s",
                e,
                exc_info=True,
            )
            return {"mensagem": t("mercado_report_error", lang=lang)}

    async def comparar_produto(
        self, produto: str, chat_id: str, lang: str = "en"
    ) -> Dict[str, Any]:
        """Compara preços entre mercados."""
        try:
            from services import get_service

            db = get_service("database").client
            res = (
                db.table("mercado_itens")
                .select("mercado, preco_unitario, data_compra, unidade")
                .eq("chat_id", chat_id)
                .ilike("produto", f"%{produto}%")
                .order("data_compra", desc=True)
                .execute()
            )
            itens = res.data or []
            if not itens:
                return {
                    "mensagem": t("mercado_no_data_product", lang=lang, product=produto)
                }

            por_mercado: Dict[str, List[float]] = {}
            for i in itens:
                p = float(i.get("preco_unitario") or 0)
                if p > 0:
                    por_mercado.setdefault(i["mercado"], []).append(p)

            if not por_mercado:
                return {
                    "mensagem": t("mercado_no_price_data", lang=lang, product=produto)
                }

            medias = {m: round(sum(ps) / len(ps), 2) for m, ps in por_mercado.items()}
            ranking = sorted(medias.items(), key=lambda x: x[1])
            mais_barato = ranking[0]
            mais_caro = ranking[-1]
            poupanca = mais_caro[1] - mais_barato[1]
            unidade = itens[0].get("unidade", "un")

            linhas = [
                t("mercado_compare_header", lang=lang, product=produto),
                "",
            ]
            for i, (mercado, preco) in enumerate(ranking):
                emoji = "✅" if i == 0 else ("❌" if i == len(ranking) - 1 else "➡️")
                n = len(por_mercado[mercado])
                linhas.append(
                    f"{emoji} *{mercado}*: "
                    f"€{preco:.2f}/{unidade} "
                    f"({t('mercado_purchase_count', lang=lang, count=n)})"
                )

            if len(ranking) > 1 and poupanca > 0.01:
                linhas += [
                    "",
                    t(
                        "mercado_compare_savings",
                        lang=lang,
                        store=mais_barato[0],
                        amount=f"{poupanca:.2f}",
                        expensive_store=mais_caro[0],
                    ),
                ]

            return {
                "mensagem": "\n".join(linhas),
                "ranking": ranking,
                "poupanca": poupanca,
            }

        except Exception as e:
            self.logger.error(
                "Erro ao comparar produto: %s",
                e,
                exc_info=True,
            )
            return {"mensagem": t("mercado_compare_error", lang=lang)}

    async def ranking_mercados(self, chat_id: str, lang: str = "en") -> Dict[str, Any]:
        """Ranking geral de mercados."""
        try:
            from services import get_service

            db = get_service("database").client
            res = (
                db.table("mercado_compras")
                .select("mercado, total")
                .eq("chat_id", chat_id)
                .execute()
            )
            compras = res.data or []
            if not compras:
                return {"mensagem": t("mercado_no_ranking_data", lang=lang)}

            totais: Dict[str, Dict] = {}
            for c in compras:
                m = c["mercado"]
                v = float(c.get("total") or 0)
                if m not in totais:
                    totais[m] = {
                        "total": 0.0,
                        "visitas": 0,
                    }
                totais[m]["total"] += v
                totais[m]["visitas"] += 1

            ranking = sorted(
                totais.items(),
                key=lambda x: x[1]["total"],
                reverse=True,
            )
            total_geral = sum(v["total"] for _, v in ranking)

            linhas = [
                t("mercado_ranking_header", lang=lang),
                "",
            ]
            medalhas = ["🥇", "🥈", "🥉"]
            for i, (mercado, dados) in enumerate(ranking):
                emoji = medalhas[i] if i < 3 else f"{i + 1}."
                pct = (dados["total"] / total_geral * 100) if total_geral else 0
                ticket = dados["total"] / dados["visitas"]
                linhas.append(
                    f"{emoji} *{mercado}*\n"
                    f" 💶 €{dados['total']:.2f} "
                    f"({pct:.0f}%) · "
                    + t("mercado_visit_count", lang=lang, count=dados["visitas"])
                    + " · "
                    + t("mercado_avg_ticket", lang=lang, amount=f"{ticket:.2f}")
                )
            if len(ranking) > 1:
                linhas += [
                    "",
                    t("mercado_cheapest_store", lang=lang, store=ranking[-1][0]),
                ]

            return {
                "mensagem": "\n".join(linhas),
                "ranking": ranking,
            }

        except Exception as e:
            self.logger.error(
                "Erro no ranking: %s",
                e,
                exc_info=True,
            )
            return {"mensagem": t("mercado_ranking_error", lang=lang)}

    async def gerar_excel_mensal(
        self, chat_id: str, mes: int = None, ano: int = None, lang: str = "en"
    ) -> Dict[str, Any]:
        """Generate monthly Excel report with all purchases."""
        import asyncio
        import io
        from datetime import datetime

        try:
            import openpyxl
            from openpyxl.styles import Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return {"sucesso": False, "mensagem": t("mercado_report_error", lang=lang)}

        now = datetime.now()
        if mes is None:
            if now.month == 1:
                mes, ano = 12, now.year - 1
            else:
                mes, ano = now.month - 1, now.year
        if ano is None:
            ano = now.year

        nome_mes = self._nome_mes(mes, lang=lang)

        # Fetch purchases
        db = self._get_client()
        start = f"{ano}-{mes:02d}-01"
        end = f"{ano + 1}-01-01" if mes == 12 else f"{ano}-{mes + 1:02d}-01"

        compras = await asyncio.get_event_loop().run_in_executor(
            None,
            lambda: (
                db.table("mercado_compras")
                .select("*")
                .eq("chat_id", chat_id)
                .gte("data_compra", start)
                .lt("data_compra", end)
                .order("data_compra")
                .execute()
            ),
        )

        if not compras.data:
            return {
                "sucesso": False,
                "mensagem": t(
                    "mercado_no_purchases_month", lang=lang, month=nome_mes, year=ano
                ),
            }

        compra_ids = [c["id"] for c in compras.data]
        itens = await asyncio.get_event_loop().run_in_executor(
            None,
            lambda: (
                db.table("mercado_itens")
                .select("*")
                .in_("compra_id", compra_ids)
                .order("compra_id,produto")
                .execute()
            ),
        )

        # ── Styles ───────────────────────────────────────────────
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(
            start_color="2E86AB", end_color="2E86AB", fill_type="solid"
        )
        currency_fmt = "€#,##0.00"
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary ─────────────────────────────────────
        ws = wb.active
        ws.title = (
            "Summary" if lang == "en" else ("Resumo" if lang == "pt" else "Resumen")
        )

        ws.merge_cells("A1:D1")
        ws["A1"].value = f"{nome_mes} {ano}"
        ws["A1"].font = Font(bold=True, size=16, color="2E86AB")

        total_gasto = sum(float(c.get("total") or 0) for c in compras.data)
        mercados = set(c.get("mercado", "") for c in compras.data)
        total_itens = len(itens.data) if itens.data else 0

        labels = {
            "en": ["Total Spent", "Stores", "Purchases", "Items"],
            "pt": ["Total Gasto", "Mercados", "Compras", "Itens"],
            "es": ["Total Gastado", "Tiendas", "Compras", "Artículos"],
        }
        lbl = labels.get(lang, labels["en"])
        values = [total_gasto, len(mercados), len(compras.data), total_itens]
        for i, (label, v) in enumerate(zip(lbl, values)):
            ws[f"A{i + 3}"] = label
            cell = ws[f"B{i + 3}"]
            cell.value = v
            if i == 0:
                cell.number_format = currency_fmt
                cell.font = Font(bold=True, size=14)

        # Spending by store table
        row = 8
        ws[f"A{row}"] = (
            "Spending by Store"
            if lang == "en"
            else ("Gastos por Mercado" if lang == "pt" else "Gastos por Tienda")
        )
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        store_headers = {
            "en": ["Store", "Total", "Visits"],
            "pt": ["Mercado", "Total", "Visitas"],
            "es": ["Tienda", "Total", "Visitas"],
        }
        for col, h in enumerate(store_headers.get(lang, store_headers["en"]), 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        gastos: Dict[str, Dict[str, Any]] = {}
        for c in compras.data:
            m = c.get("mercado", "?")
            if m not in gastos:
                gastos[m] = {"total": 0, "visitas": 0}
            gastos[m]["total"] += float(c.get("total") or 0)
            gastos[m]["visitas"] += 1

        for m, d in sorted(gastos.items(), key=lambda x: -x[1]["total"]):
            ws.cell(row=row, column=1, value=m).border = thin_border
            c_cell = ws.cell(row=row, column=2, value=d["total"])
            c_cell.number_format = currency_fmt
            c_cell.border = thin_border
            ws.cell(row=row, column=3, value=d["visitas"]).border = thin_border
            row += 1

        for col_idx in range(1, 5):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        # ── Sheet 2: Items ───────────────────────────────────────
        ws2 = wb.create_sheet(title="Items" if lang == "en" else "Itens")
        item_headers = {
            "en": ["Date", "Store", "Product", "Price", "Category"],
            "pt": ["Data", "Mercado", "Produto", "Preço", "Categoria"],
            "es": ["Fecha", "Tienda", "Producto", "Precio", "Categoría"],
        }
        for col, h in enumerate(item_headers.get(lang, item_headers["en"]), 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        compra_map = {c["id"]: c for c in compras.data}
        row = 2
        for item in itens.data or []:
            compra = compra_map.get(item.get("compra_id"), {})
            ws2.cell(
                row=row, column=1, value=compra.get("data_compra", "")
            ).border = thin_border
            ws2.cell(
                row=row, column=2, value=compra.get("mercado", "")
            ).border = thin_border
            ws2.cell(
                row=row, column=3, value=item.get("produto", "")
            ).border = thin_border
            price_cell = ws2.cell(
                row=row, column=4, value=float(item.get("preco") or 0)
            )
            price_cell.number_format = currency_fmt
            price_cell.border = thin_border
            cat_pt = item.get("categoria", "Outros")
            cat_key = _CAT_I18N.get(cat_pt, "cat_other")
            ws2.cell(
                row=row, column=5, value=t(cat_key, lang=lang)
            ).border = thin_border
            row += 1

        for col_idx in range(1, 6):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 18

        # ── Sheet 3: Price Comparison (if 2+ stores) ─────────────
        if len(mercados) > 1:
            ws3 = wb.create_sheet(
                title="Price Compare" if lang == "en" else "Comparação"
            )
            cmp_headers = {
                "en": [
                    "Product",
                    "Cheapest",
                    "Min Price",
                    "Expensive",
                    "Max Price",
                    "Savings",
                ],
                "pt": [
                    "Produto",
                    "Mais Barato",
                    "Menor Preço",
                    "Mais Caro",
                    "Maior Preço",
                    "Poupança",
                ],
                "es": [
                    "Producto",
                    "Más Barato",
                    "Precio Min",
                    "Más Caro",
                    "Precio Max",
                    "Ahorro",
                ],
            }
            for col, h in enumerate(cmp_headers.get(lang, cmp_headers["en"]), 1):
                cell = ws3.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            price_map: Dict[str, Dict[str, float]] = {}
            for item in itens.data or []:
                compra = compra_map.get(item.get("compra_id"), {})
                prod = item.get("produto", "").lower()
                store = compra.get("mercado", "")
                price = float(item.get("preco") or 0)
                if prod and store and price > 0:
                    if prod not in price_map:
                        price_map[prod] = {}
                    if store not in price_map[prod] or price < price_map[prod][store]:
                        price_map[prod][store] = price

            row = 2
            for prod, stores in sorted(price_map.items()):
                if len(stores) < 2:
                    continue
                cheapest = min(stores, key=stores.get)
                expensive = max(stores, key=stores.get)
                savings = stores[expensive] - stores[cheapest]
                ws3.cell(row=row, column=1, value=prod.title()).border = thin_border
                ws3.cell(row=row, column=2, value=cheapest).border = thin_border
                min_cell = ws3.cell(row=row, column=3, value=stores[cheapest])
                min_cell.number_format = currency_fmt
                min_cell.border = thin_border
                ws3.cell(row=row, column=4, value=expensive).border = thin_border
                max_cell = ws3.cell(row=row, column=5, value=stores[expensive])
                max_cell.number_format = currency_fmt
                max_cell.border = thin_border
                sav_cell = ws3.cell(row=row, column=6, value=savings)
                sav_cell.number_format = currency_fmt
                sav_cell.border = thin_border
                sav_cell.font = Font(color="008000", bold=True)
                row += 1

            for col_idx in range(1, 7):
                ws3.column_dimensions[get_column_letter(col_idx)].width = 20

        # ── Save to bytes ────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        filename = f"mercado_report_{ano}_{mes:02d}.xlsx"
        return {
            "sucesso": True,
            "excel_bytes": excel_bytes,
            "filename": filename,
            "resumo": f"{nome_mes} {ano}: €{total_gasto:.2f} | {len(compras.data)} purchases | {total_itens} items | {len(mercados)} stores",
        }

    def _nome_mes(self, mes: int, lang: str = "en") -> str:
        """Get localized month name."""
        return t(f"month_{mes}", lang=lang) if 1 <= mes <= 12 else str(mes)

    # ------------------------------------------------------------------
    # Smart Shopping Reminder (daily check for idle lists)
    # ------------------------------------------------------------------

    async def verificar_lista_pendente(
        self, chat_id: str, lang: str = "en"
    ) -> Dict[str, Any]:
        """
        Check if user has a shopping list that's been idle for 3+ days.

        Returns:
            Dict with:
            - has_pending: bool
            - item_count: int
            - items_preview: str (first 5 items, comma separated)
            - days_old: int (days since oldest item was added)
            - mensagem: str (formatted message)
        """
        try:
            import asyncio

            from services import get_service

            db = get_service("database").client

            res = await asyncio.get_event_loop().run_in_executor(
                None,
                lambda: (
                    db.table("shopping_list")
                    .select("item, created_at")
                    .eq("user_id", chat_id)
                    .order("created_at", desc=False)
                    .execute()
                ),
            )

            items = res.data or []
            if not items:
                return {"has_pending": False, "item_count": 0, "mensagem": ""}

            # Calculate days since oldest item
            oldest_str = items[0].get("created_at", "")
            if oldest_str:
                try:
                    oldest = datetime.fromisoformat(
                        oldest_str.replace("Z", "+00:00")
                    ).date()
                    days_old = (date.today() - oldest).days
                except (ValueError, TypeError):
                    days_old = 0
            else:
                days_old = 0

            if days_old < 3:
                return {"has_pending": False, "item_count": len(items), "mensagem": ""}

            # Build preview (first 5 items)
            item_names = [i.get("item", "") for i in items if i.get("item")]
            preview = ", ".join(item_names[:5])
            if len(item_names) > 5:
                preview += f" (+{len(item_names) - 5})"

            msg = t(
                "mercado_shopping_reminder",
                lang=lang,
                count=len(item_names),
                days=days_old,
                items=preview,
            )

            return {
                "has_pending": True,
                "item_count": len(item_names),
                "items_preview": preview,
                "days_old": days_old,
                "mensagem": msg,
            }

        except Exception as e:
            self.logger.error("Shopping reminder check failed: %s", e, exc_info=True)
            return {"has_pending": False, "item_count": 0, "mensagem": ""}
